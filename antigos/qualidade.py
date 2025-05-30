#https://www.youtube.com/watch?v=bu5wXjz2KvU

import pandas as pd
import numpy as np
import os
import datetime 
import gspread
import streamlit as st
import time
import zipfile

from datetime import datetime
from datetime import timedelta
from pathlib import Path
from openpyxl import Workbook, load_workbook
from PIL import Image
from openpyxl.drawing.image import Image as xlImage
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from google.oauth2 import service_account
from PIL import Image as PILImage

# data = '06/02/2024'
###### CONECTANDO PLANILHAS ##########

st.title('Checklist Qualidade')

name_sheet = 'RQ AV-002-000 (PLANILHA DE CARGAS)'
name_sheet2 = 'Cópia de RQ CQ-011-000 (Checklist da Qualidade)'

worksheet1 = 'Importar Dados'
worksheet2 = 'Dados'
# worksheet3 = 'Cópia de RQ CQ-011-000 (Checklist da Qualidade)'
worksheet4 = 'RODAS E CILINDROS'

#filename = r"C:\Users\pcp2\ordem de producao\Ordem-de-producao\service_account.json"
filename = "service_account.json"

service_account_info = st.secrets["GOOGLE_SERVICE_ACCOUNT"]

scope = ['https://www.googleapis.com/auth/spreadsheets',
         "https://www.googleapis.com/auth/drive"]

credentials = service_account.Credentials.from_service_account_info(service_account_info, scopes=scope)

sa = gspread.authorize(credentials)
sh = sa.open(name_sheet)
sh2 = sa.open(name_sheet2)

wks1 = sh.worksheet(worksheet1)
wks2 = sh2.worksheet(worksheet2)
# wks3 = sh2.worksheet(worksheet3)
wks4 = sh2.worksheet(worksheet4)

#obtendo todos os valores da planilha
list1 = wks1.get()
list2 = wks2.get()
# list3 = wks3.get()
list4 = wks4.get()

#transformando em dataframe
importar_dados = pd.DataFrame(list1)
dados = pd.DataFrame(list2[1:],columns=list2[0])
# checklist_qualidade = pd.DataFrame(list3)

colunas = ['carreta','CÓDIGO','descrição','214105','214104','214107','214108','268502','033703','034478','033516','214114','262728','034149','035003','225679','035348','035390','036048','240471','240474','240485','240640','240643','222416','036117','034550','034630','034830','RODA','QUANTIDADE1','CILINDRO','QUANTIDADE2']
colunas2 = ['codigo','descricao']

rodas_cilindros = pd.DataFrame(list4).iloc[:,0:33]
rodas_cilindros = rodas_cilindros.set_axis(colunas,axis=1)

codigos_descricao = pd.DataFrame(list4).iloc[:,29:31]
codigos_descricao = codigos_descricao.set_axis(colunas2,axis=1)
codigos_descricao = codigos_descricao.dropna()

###### TRATANDO DADOS #########

#####Tratando datas######

importar_dados = importar_dados[[1,3,6,9,10,11,13]]
importar_dados = importar_dados.rename(columns={1:'Estado',3:'Cidade',6: 'Datas',9:'Pessoa',10:'Recurso',11:'Descrição',13:'Serie'})
importar_dados['Cidade/Estado'] = importar_dados['Cidade'] + '/' + importar_dados['Estado']
importar_dados.drop(columns=['Estado', 'Cidade'], inplace=True)
importar_dados = importar_dados[1:]
importar_dados = importar_dados[importar_dados['Datas'].notnull() & (importar_dados['Datas'] != '')]

#####Valores nulos######

importar_dados.dropna(inplace=True)
importar_dados = importar_dados.reset_index(drop=True)

dados = dados[dados['REF PRINCIPAL'].notnull() & (dados['REF PRINCIPAL'] != '')]
dados = dados.reset_index(drop=True)

today = datetime.now()
ts = pd.Timestamp(today)
today = today.strftime('%d/%m/%Y')

filenames=[]

with st.sidebar:
    img_path = 'logo-cemagL.png'
    pil_img = PILImage.open(img_path)
    st.image(pil_img, width=300)

with st.form(key='my_form'):
    
    with st.sidebar:
        
        tipo_filtro = st.date_input('Data: ')
        tipo_filtro = tipo_filtro.strftime("%d/%m/%Y")
        # tipo_filtro = "17/02/2025"
        submit_button = st.form_submit_button(label='Gerar')

if submit_button:

    # importar_dados['Datas'] = pd.to_datetime(importar_dados['Datas'], format="%d/%m/%Y").dt.strftime("%d/%m/%Y")
    importar_dados['Datas'] = pd.to_datetime(importar_dados['Datas'], format="%d/%m/%Y", errors='coerce')
    importar_dados['Datas'] = importar_dados['Datas'].dt.strftime("%d/%m/%Y")

    dados_filtrados = importar_dados[importar_dados['Datas'] == tipo_filtro]
    dados_filtrados = dados_filtrados[dados_filtrados['Serie'].notnull() & (dados_filtrados['Serie'] != '')]
    dados_filtrados.reset_index(drop=True,inplace=True)

    df_dados = dados.copy()

    dados_filtrados['cor'] = ''

    df_cores = pd.DataFrame({'Recurso_cor':['AN','VJ','LC','VM','AV','sem_cor'], 
        'cor':['Azul','Verde','Laranja','Vermelho','Amarelo','Laranja']})

    for i in range(len(dados_filtrados)):
        
        inicio = len(dados_filtrados['Recurso'][i])-2
        fim = len(dados_filtrados['Recurso'][i])

        dados_filtrados['cor'][i] = dados_filtrados['Recurso'][i][inicio:fim]
        
        df_cores_filtrada = df_cores[df_cores['Recurso_cor'] == dados_filtrados['cor'][i]].reset_index(drop=True)

        if len(df_cores_filtrada) == 0:
            dados_filtrados['cor'][i] = 'Laranja'
        else:
            cor = df_cores_filtrada['cor'][0]
            dados_filtrados['cor'][i] = cor 

    dados_filtrados['Recurso']=dados_filtrados['Recurso'].str.replace(' AN','') # Azul
    dados_filtrados['Recurso']=dados_filtrados['Recurso'].str.replace(' VJ','') # Verde
    dados_filtrados['Recurso']=dados_filtrados['Recurso'].str.replace(' LC','') # Laranja
    dados_filtrados['Recurso']=dados_filtrados['Recurso'].str.replace(' VM','') # Vermelho
    dados_filtrados['Recurso']=dados_filtrados['Recurso'].str.replace(' AV','') # Amarelo
    
    for i in range(len(dados_filtrados)):
        
        nome_carreta = dados_filtrados['Recurso'][i]
        descricao_carreta = dados_filtrados['Descrição'][i]

        codigo_descricao = str(nome_carreta) + str(descricao_carreta)

        wb = Workbook()
        # wb = load_workbook('modelo_op_carpintaria1.xlsx')
        wb = load_workbook('modelo_expedicao_checklist2307.xlsx')
        ws = wb.active

        ws['B7'] = dados_filtrados['Recurso'][i] 
        ws['A8'] = dados_filtrados['Descrição'][i] 
        ws['G8'] = dados_filtrados['Pessoa'][i] #  data de hoje
        ws['B9'] = dados_filtrados['Serie'][i]
        ws['G9'] = dados_filtrados['Cidade/Estado'][i]
        ws['C9']  = tipo_filtro  # data da carga
        ws['F7'] = dados_filtrados['cor'][i]
        ws['A36'] = 'ACESSÓRIOS 01'
        ws['B36'] = 'ACESSÓRIOS'
        ws['F36'] = 1

        df_filtrado = df_dados[df_dados['REF PRINCIPAL'] == nome_carreta]
        df_filtrado = df_filtrado.reset_index(drop=True)
        
        if codigo_descricao.lower().find('bb') != -1:
            
            # Se tiver bomba dentro do código
            ws['A32'] = '200391'
            ws['B32'] = 'BOMBA'
            ws['F32'] = 1
                
        if codigo_descricao.lower().find('rs/rd') != -1: 
            
            # Se tiver rs/rd dentro do código
            ws['A31'] = '214108'
            ws['B31'] = 'RODA 6 FUROS TANDEM FA6 Flag Romaneio'
            ws['F31'] = 2

        if codigo_descricao.lower().find('rs/rd') != -1 and codigo_descricao.lower().find('roda 20') != -1: 
            
            # Se tiver rs/rd dentro do código e aro 20
            ws['A32'] = ''
            ws['B32'] = ''
            ws['F32'] = ''

            ws['A31'] = '035390'
            ws['B31'] = 'RODA RS R20 CINZA [CBH12/FT12500] Flag Romaneio'
            ws['F31'] = 2

        if codigo_descricao.lower().find('rs/rd') != -1 and (codigo_descricao.lower().find('(i)') != -1 or codigo_descricao.lower().find('(r)') != -1):
            
            # Se tiver rs/rd dentro do código
            ws['A31'] = '214108'
            ws['B31'] = 'RODA 6 FUROS TANDEM FA6 Flag Romaneio'
            ws['F31'] = 2

            # Se tiver pneu dentro do código
            ws['A32'] = 'Pneus'
            ws['B32'] = 'PNEUS'
            ws['F32'] = 6

        if not codigo_descricao.lower().find('rs/rd') != -1 and (codigo_descricao.lower().find('(i)') != -1 or codigo_descricao.lower().find('(r)') != -1):

            # Se tiver rs/rd dentro do código
            # ws['A42'] = '214108'
            # ws['B42'] = 'RODA 6 FUROS TANDEM FA6 Flag Romaneio'
            # ws['D42'] = 2

            # Se tiver pneu dentro do código
            ws['A32'] = 'Pneus'
            ws['B32'] = 'PNEUS'
            ws['F32'] = 4

        #Tirante
        if codigo_descricao.lower().find('roçax') != -1 or codigo_descricao.lower().find('f6') != -1 or codigo_descricao.lower().find('f4') != -1 or codigo_descricao.lower().find('fa5') != -1 or codigo_descricao.lower().find('ftc4300') != -1 or codigo_descricao.lower().find('ftc6500') != -1 or codigo_descricao.lower().find('ftc10500') != -1: 
            ws['A33'] = ''
            ws['B33'] = ''
            ws['F33'] = ''
        else:
            ws['A33'] = 'TIRANTE DA TRAVA COMPLETO'
            ws['B33'] = 'TIRANTE DA TRAVA COMPLETO'
            ws['F33'] = 2

        filtro_rodas_cilindros = rodas_cilindros[rodas_cilindros['carreta'] == nome_carreta]

        if len(filtro_rodas_cilindros) > 0:
            # Verificação se existe roda/cilindo para essa carreta
            filtro_rodas_cilindros.reset_index(inplace=True, drop=True)
            
            try:
                # Rodas

                codigos_descricao_roda = codigos_descricao[codigos_descricao['codigo'] == filtro_rodas_cilindros['RODA'][0]].reset_index(drop=True)['descricao'][0]
                ws['A37'] = filtro_rodas_cilindros['RODA'][0]
                ws['B37'] = '' #sem descrição
                ws['F37'] = filtro_rodas_cilindros['QUANTIDADE1'][0]
            except:
                ws['A37'] = ''
                ws['B37'] = ''
                ws['F37'] = ''

            try:
                # Cilindros
            
                codigos_descricao_cilindro = codigos_descricao[codigos_descricao['codigo'] == filtro_rodas_cilindros['CILINDRO'][0]].reset_index(drop=True)['descricao'][0]
                ws['A34'] = filtro_rodas_cilindros['CILINDRO'][0]
                ws['B34'] = codigos_descricao_cilindro
                ws['F34'] = filtro_rodas_cilindros['QUANTIDADE2'][0]
                
            except:
                ws['A34'] = ''
                ws['B34'] = ''
                ws['F34'] = ''

        j = 12
        for k in range(len(df_filtrado)):
        
            if j >= 29:
                ws.insert_rows(j)  # insere nova linha antes de usar a linha j

            ws['A' + str(j)] = df_filtrado['Recurso'][k] 
            ws['B' + str(j)] = df_filtrado['Descrição'][k] 
            ws['C' + str(j)] = df_filtrado['qnt'][k] 
            j+=1
                    
        wb.template = False
        wb.save("Arquivo" + str(i) +'.xlsx') 
        my_file = "Arquivo" + str(i) +'.xlsx'
        filenames.append(my_file)

    filenames_unique = list(set(filenames))

    with zipfile.ZipFile("Arquivos.zip", mode="w") as archive:
        for filename in filenames_unique:
            archive.write(filename)

    with open("Arquivos.zip", "rb") as fp:
        btn = st.download_button(
            label="Download arquivos",
            data=fp,
            file_name="Arquivos.zip",
            mime="application/zip"
        )